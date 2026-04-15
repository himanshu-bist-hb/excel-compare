"""
Excel Comparator Pro
A professional Streamlit app to compare two Excel files at sheet and cell level.
"""

import copy
import io
import re
from datetime import datetime
from typing import Dict, List, Set, Tuple

import numpy as np
import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.properties import PageSetupProperties

# ─────────────────────────────────────────────────────────────────────────────
# Page Config
# ─────────────────────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Excel Comparator Pro",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ─────────────────────────────────────────────────────────────────────────────
# CSS
# ─────────────────────────────────────────────────────────────────────────────

CUSTOM_CSS = """
<style>
  html, body, [class*="css"] {
    font-family: 'Segoe UI', system-ui, -apple-system, sans-serif;
  }

  /* ── Hero banner ── */
  .hero {
    background: linear-gradient(135deg, #0f2942 0%, #1a4a7a 60%, #1e6091 100%);
    color: white;
    padding: 2.4rem 2rem 2rem;
    border-radius: 14px;
    margin-bottom: 1.8rem;
    text-align: center;
    box-shadow: 0 8px 32px rgba(15,41,66,0.28);
  }
  .hero h1 { margin:0; font-size:2.3rem; font-weight:800; letter-spacing:-0.5px; }
  .hero p  { margin:0.5rem 0 0; opacity:0.82; font-size:1.05rem; }

  /* ── Metric cards ── */
  .metric-grid { display:flex; gap:14px; flex-wrap:wrap; margin:1rem 0; }
  .metric-card {
    flex:1 1 130px;
    background:white;
    border-radius:12px;
    padding:1.1rem 1rem 0.9rem;
    border:1px solid #e4e8ef;
    text-align:center;
    box-shadow:0 2px 8px rgba(0,0,0,0.05);
  }
  .metric-value { font-size:2rem; font-weight:800; line-height:1.1; }
  .metric-label { font-size:11px; color:#6b7280; text-transform:uppercase;
                  letter-spacing:0.6px; margin-top:4px; }

  /* ── Sheet pill tags ── */
  .sheet-pills {
    display:flex; flex-wrap:wrap; gap:8px;
    padding:1rem 1.1rem; background:#f4f6f9;
    border-radius:10px; margin:0.8rem 0 1.2rem;
  }
  .pill {
    padding:5px 15px; border-radius:20px; font-size:13px;
    font-weight:600; border:2px solid; cursor:default;
    display:inline-flex; align-items:center; gap:5px;
  }
  .pill-unchanged { background:#edf0f4; color:#374151; border-color:#c5cdd8; }
  .pill-new       { background:#d1fae5; color:#065f46; border-color:#10b981; }
  .pill-deleted   { background:#fee2e2; color:#991b1b; border-color:#ef4444; }
  .pill-modified  { background:#fef3c7; color:#78350f; border-color:#f59e0b; }

  /* ── Legend ── */
  .legend { display:flex; flex-wrap:wrap; gap:18px; margin:0.6rem 0 1rem; }
  .legend-item { display:flex; align-items:center; gap:7px; font-size:13px; color:#374151; }
  .legend-dot  { width:15px; height:15px; border-radius:4px; flex-shrink:0; }

  /* ── Diff table ── */
  .diff-wrap {
    overflow-x:auto;
    border-radius:10px;
    border:1px solid #e4e8ef;
    box-shadow:0 2px 8px rgba(0,0,0,0.04);
    max-height:520px;
    overflow-y:auto;
  }
  table.diff {
    border-collapse:collapse;
    width:100%;
    font-size:13px;
    white-space:nowrap;
  }
  table.diff thead th {
    background:#0f2942;
    color:white;
    padding:9px 14px;
    font-weight:600;
    text-align:left;
    position:sticky;
    top:0;
    z-index:2;
    border-right:1px solid rgba(255,255,255,0.1);
  }
  table.diff thead th:first-child { width:42px; text-align:center; }
  table.diff tbody td {
    padding:7px 14px;
    border-bottom:1px solid #f0f2f5;
    border-right:1px solid #f0f2f5;
    vertical-align:top;
    max-width:260px;
    overflow:hidden;
    text-overflow:ellipsis;
  }
  table.diff tbody tr:hover td { filter:brightness(0.97); }

  /* row-level colours */
  .r-added   td { background:#ecfdf5 !important; }
  .r-deleted td { background:#fef2f2 !important; }

  /* cell-level colours */
  .c-changed { background:#fffbeb !important; }
  .c-added   { background:#ecfdf5 !important; }
  .c-deleted { background:#fef2f2 !important; }

  .rn { color:#9ca3af; font-size:11px; text-align:center; user-select:none; }

  /* inline old→new diff inside a changed cell */
  .val-old  { text-decoration:line-through; color:#dc2626; font-size:11px;
              display:block; line-height:1.3; }
  .val-new  { color:#16a34a; font-size:12px; display:block;
              font-weight:600; line-height:1.4; }
  .val-only { font-size:13px; }

  /* ── Info / instruction box ── */
  .info-box {
    background:#eff6ff; border-left:4px solid #3b82f6;
    padding:1rem 1.2rem; border-radius:0 10px 10px 0; margin:0.5rem 0;
    font-size:14px; line-height:1.7; color:#1e3a5f;
  }
  .info-box ol, .info-box ul { margin:0.4rem 0 0; padding-left:1.4rem; }
  .info-box strong { color:#0f2942; }

  /* ── Upload labels ── */
  .upload-label {
    font-size:15px; font-weight:700; color:#0f2942;
    margin-bottom:0.3rem; display:block;
  }

  /* ── Misc ── */
  .divider { border:none; border-top:1.5px solid #e4e8ef; margin:1.6rem 0; }
  .section-title { font-size:1.1rem; font-weight:700; color:#0f2942; margin:0 0 0.6rem; }
  .truncation-note {
    text-align:center; padding:0.8rem 1rem; color:#6b7280;
    font-style:italic; font-size:13px; background:#f9fafb;
  }

  /* hide streamlit default branding */
  #MainMenu { visibility:hidden; }
  footer    { visibility:hidden; }
  .block-container { padding-top:1.5rem; }
</style>
"""

st.markdown(CUSTOM_CSS, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# Data helpers
# ─────────────────────────────────────────────────────────────────────────────

def _file_bytes(uploaded_file) -> bytes:
    """Read bytes from a Streamlit UploadedFile safely."""
    if hasattr(uploaded_file, "seek"):
        uploaded_file.seek(0)
    return uploaded_file.read()


def _detect_engine(raw: bytes, name: str) -> str:
    """Pick openpyxl or xlrd based on file extension."""
    return "xlrd" if name.lower().endswith(".xls") else "openpyxl"


def read_excel_sheets(raw: bytes, filename: str) -> Dict[str, pd.DataFrame]:
    """Load every sheet of an Excel file into {sheet_name: DataFrame}."""
    engine = _detect_engine(raw, filename)
    try:
        buf = io.BytesIO(raw)
        xl = pd.ExcelFile(buf, engine=engine)
        result: Dict[str, pd.DataFrame] = {}
        for sheet in xl.sheet_names:
            buf.seek(0)
            df = pd.read_excel(
                buf, sheet_name=sheet, header=None,
                dtype=str, engine=engine
            )
            df = df.fillna("")
            result[sheet] = df
        return result
    except Exception as exc:
        st.error(f"Could not read **{filename}**: {exc}")
        return {}


def cell_str(val) -> str:
    """Normalise a cell value to a clean comparable string."""
    s = str(val).strip() if val is not None else ""
    # strip trailing .0 produced by float-to-str conversion
    if re.fullmatch(r"-?\d+\.0+", s):
        s = s[: s.index(".")]
    return s


def compare_dataframes(
    old_df: pd.DataFrame, new_df: pd.DataFrame
) -> Tuple[pd.DataFrame, pd.DataFrame, dict, dict, dict]:
    """
    Align two DataFrames and return per-cell / per-row change status.

    Returns
    -------
    old_aligned, new_aligned : padded to same shape
    cell_status  : {(row, col): 'same' | 'changed' | 'added' | 'deleted'}
    row_status   : {row: 'same' | 'changed' | 'added' | 'deleted'}
    stats        : summary counts
    """
    nr = max(len(old_df), len(new_df))
    nc = max(
        len(old_df.columns) if len(old_df) else 0,
        len(new_df.columns) if len(new_df) else 0,
    )
    if nc == 0:
        empty = pd.DataFrame()
        return empty, empty, {}, {}, {
            "total_rows": 0, "total_cols": 0,
            "added_rows": 0, "deleted_rows": 0,
            "changed_rows": 0, "changed_cells": 0,
        }

    old_a = old_df.reindex(range(nr)).reindex(columns=range(nc)).fillna("")
    new_a = new_df.reindex(range(nr)).reindex(columns=range(nc)).fillna("")

    old_row_range = set(range(len(old_df)))
    new_row_range = set(range(len(new_df)))

    cell_status: dict = {}
    row_status: dict = {}
    added_rows = deleted_rows = changed_rows = changed_cells = 0

    for i in range(nr):
        in_old = i in old_row_range
        in_new = i in new_row_range

        if in_new and not in_old:
            row_status[i] = "added"
            added_rows += 1
            for j in range(nc):
                cell_status[(i, j)] = "added"

        elif in_old and not in_new:
            row_status[i] = "deleted"
            deleted_rows += 1
            for j in range(nc):
                cell_status[(i, j)] = "deleted"

        else:
            row_changed = False
            for j in range(nc):
                ov = cell_str(old_a.iat[i, j])
                nv = cell_str(new_a.iat[i, j])
                if ov != nv:
                    cell_status[(i, j)] = "changed"
                    row_changed = True
                    changed_cells += 1
                else:
                    cell_status[(i, j)] = "same"
            if row_changed:
                row_status[i] = "changed"
                changed_rows += 1
            else:
                row_status[i] = "same"

    return old_a, new_a, cell_status, row_status, {
        "total_rows":    nr,
        "total_cols":    nc,
        "added_rows":    added_rows,
        "deleted_rows":  deleted_rows,
        "changed_rows":  changed_rows,
        "changed_cells": changed_cells,
    }


# ─────────────────────────────────────────────────────────────────────────────
# HTML rendering
# ─────────────────────────────────────────────────────────────────────────────

_MAX_DISPLAY_ROWS = 1000


def _esc(text: str) -> str:
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def render_diff_table(
    old_a: pd.DataFrame,
    new_a: pd.DataFrame,
    cell_status: dict,
    row_status: dict,
    max_rows: int = _MAX_DISPLAY_ROWS,
) -> str:
    """Return an HTML string of the colour-coded diff table."""
    nr = len(old_a)
    nc = len(old_a.columns)
    col_headers = [get_column_letter(j + 1) for j in range(nc)]

    parts = ['<div class="diff-wrap"><table class="diff"><thead><tr>']
    parts.append('<th class="rn">#</th>')
    for ch in col_headers:
        parts.append(f"<th>{ch}</th>")
    parts.append("</tr></thead><tbody>")

    show = min(nr, max_rows)
    for i in range(show):
        rs = row_status.get(i, "same")
        row_cls = {"added": "r-added", "deleted": "r-deleted"}.get(rs, "")
        parts.append(f'<tr class="{row_cls}">')
        parts.append(f'<td class="rn">{i + 1}</td>')

        for j in range(nc):
            cs = cell_status.get((i, j), "same")
            ov = _esc(cell_str(old_a.iat[i, j]))
            nv = _esc(cell_str(new_a.iat[i, j]))

            if rs == "added":
                cell_cls = "c-added"
                inner = f'<span class="val-only">{nv}</span>'
            elif rs == "deleted":
                cell_cls = "c-deleted"
                inner = f'<span class="val-only">{ov}</span>'
            elif cs == "changed":
                cell_cls = "c-changed"
                if ov and nv:
                    inner = (
                        f'<span class="val-old">{ov}</span>'
                        f'<span class="val-new">{nv}</span>'
                    )
                else:
                    inner = f'<span class="val-only">{nv or ov}</span>'
            else:
                cell_cls = ""
                inner = f'<span class="val-only">{nv}</span>'

            td_cls = f' class="{cell_cls}"' if cell_cls else ""
            parts.append(f"<td{td_cls}>{inner}</td>")

        parts.append("</tr>")

    if nr > max_rows:
        parts.append(
            f'<tr><td colspan="{nc + 1}" class="truncation-note">'
            f"⚠ Showing first {max_rows:,} of {nr:,} rows. "
            f"Download the highlighted Excel report to view all rows.</td></tr>"
        )

    parts.append("</tbody></table></div>")
    return "".join(parts)


# ─────────────────────────────────────────────────────────────────────────────
# Excel export
# ─────────────────────────────────────────────────────────────────────────────

# Colour fills
_FILL_CHANGED  = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
_FILL_ADDED    = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
_FILL_DELETED  = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
_FILL_HEADER   = PatternFill(start_color="0F2942", end_color="0F2942", fill_type="solid")
_FILL_SUMMARY_HEADER = PatternFill(start_color="1A4A7A", end_color="1A4A7A", fill_type="solid")
_FONT_HEADER   = Font(color="FFFFFF", bold=True, name="Segoe UI", size=10)
_FONT_NORMAL   = Font(name="Segoe UI", size=10)
_FONT_BOLD     = Font(name="Segoe UI", size=10, bold=True)

_TAB_COLOR = {
    "new":       "10B981",
    "deleted":   "EF4444",
    "modified":  "F59E0B",
    "unchanged": "6B7280",
}

_THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

# ── Side-by-side view fills / fonts / borders ────────────────────────────────
_FILL_SBS_CHG_NEW   = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
_FILL_SBS_NEW_SHEET = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
_FILL_SBS_DEL_SHEET = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
_FILL_SBS_SEP_DATA  = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_FILL_SBS_HDR_OLD   = PatternFill(start_color="0F2942", end_color="0F2942", fill_type="solid")
_FILL_SBS_HDR_NEW   = PatternFill(start_color="1E6091", end_color="1E6091", fill_type="solid")
_FILL_SBS_HDR_SEP   = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
_FONT_STRIKE        = Font(name="Segoe UI", size=10, strike=True, color="C00000")
_FONT_SBS_HDR       = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
_BORDER_SEP_DATA    = Border(
    left=Side(style="medium", color="595959"),
    right=Side(style="medium", color="595959"),
    top=Side(style="thin",   color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def _auto_width(ws):
    """Auto-fit column widths (capped at 60)."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def _apply_border(ws, max_row: int, max_col: int):
    """Apply thin borders to the data range."""
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(r, c).border = _THIN_BORDER


def build_summary_sheet(
    wb: openpyxl.Workbook,
    ordered: List[str],
    new_only: Set[str],
    deleted_only: Set[str],
    sheet_stats: Dict[str, dict],
    old_filename: str,
    new_filename: str,
):
    """Create a formatted Summary sheet at the front of the workbook."""
    ws = wb.create_sheet("📋 Summary", 0)
    ws.sheet_properties.tabColor = "0F2942"

    # Title block
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "Excel Comparison Report"
    title_cell.font = Font(name="Segoe UI", size=16, bold=True, color="0F2942")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Meta rows
    meta = [
        ("Original File",  old_filename),
        ("Revised File",   new_filename),
        ("Generated",      datetime.now().strftime("%Y-%m-%d  %H:%M:%S")),
    ]
    for idx, (label, value) in enumerate(meta, start=2):
        ws.cell(idx, 1, label).font = _FONT_BOLD
        ws.cell(idx, 2, value).font = _FONT_NORMAL
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 18

    # Column headers for sheet table
    headers = ["Sheet Name", "Status", "Changed Cells",
               "Added Rows", "Deleted Rows", "Changed Rows", "Notes"]
    header_row = 6
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(header_row, c, h)
        cell.fill = _FILL_SUMMARY_HEADER
        cell.font = _FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[header_row].height = 22

    STATUS_FILL = {
        "Added":     PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        "Deleted":   PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        "Modified":  PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        "Unchanged": PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"),
    }

    for row_off, sname in enumerate(ordered, start=1):
        r = header_row + row_off
        if sname in new_only:
            status, sv = "Added", {}
            note = "New sheet in revised file"
        elif sname in deleted_only:
            status, sv = "Deleted", {}
            note = "Removed from revised file"
        else:
            sv = sheet_stats.get(sname, {})
            has_chg = sv.get("changed_cells", 0) + sv.get("added_rows", 0) + sv.get("deleted_rows", 0) > 0
            status = "Modified" if has_chg else "Unchanged"
            note = ""

        row_fill = STATUS_FILL.get(status)
        data = [
            sname,
            status,
            sv.get("changed_cells", "—"),
            sv.get("added_rows",    "—"),
            sv.get("deleted_rows",  "—"),
            sv.get("changed_rows",  "—"),
            note,
        ]
        for c, val in enumerate(data, start=1):
            cell = ws.cell(r, c, val)
            cell.font = _FONT_NORMAL
            cell.alignment = Alignment(vertical="center")
            if row_fill:
                cell.fill = row_fill

    _apply_border(ws, header_row + len(ordered), len(headers))
    _auto_width(ws)


def build_highlighted_excel(
    old_sheets: Dict[str, pd.DataFrame],
    new_sheets: Dict[str, pd.DataFrame],
    ordered: List[str],
    new_only: Set[str],
    deleted_only: Set[str],
    sheet_stats: Dict[str, dict],
    old_filename: str,
    new_filename: str,
) -> bytes:
    """Build a complete highlighted Excel workbook and return its bytes."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Data sheets ──────────────────────────────────────────────────────────
    for name in ordered:
        safe_name = name[:31]
        ws = wb.create_sheet(title=safe_name)

        if name in new_only:
            ws.sheet_properties.tabColor = _TAB_COLOR["new"]
            df = new_sheets.get(name, pd.DataFrame())
            if not df.empty:
                for i, row in enumerate(dataframe_to_rows(df, index=False, header=False), 1):
                    for j, val in enumerate(row, 1):
                        c = ws.cell(i, j, None if val == "" else val)
                        c.fill = _FILL_ADDED
                        c.font = _FONT_NORMAL
                _apply_border(ws, len(df), len(df.columns))
                _auto_width(ws)

        elif name in deleted_only:
            ws.sheet_properties.tabColor = _TAB_COLOR["deleted"]
            df = old_sheets.get(name, pd.DataFrame())
            if not df.empty:
                for i, row in enumerate(dataframe_to_rows(df, index=False, header=False), 1):
                    for j, val in enumerate(row, 1):
                        c = ws.cell(i, j, None if val == "" else val)
                        c.fill = _FILL_DELETED
                        c.font = _FONT_NORMAL
                _apply_border(ws, len(df), len(df.columns))
                _auto_width(ws)

        else:
            old_df = old_sheets.get(name, pd.DataFrame())
            new_df = new_sheets.get(name, pd.DataFrame())
            old_a, new_a, cell_status, row_status, stats = compare_dataframes(old_df, new_df)

            has_chg = stats["changed_cells"] + stats["added_rows"] + stats["deleted_rows"] > 0
            ws.sheet_properties.tabColor = (
                _TAB_COLOR["modified"] if has_chg else _TAB_COLOR["unchanged"]
            )

            nr, nc = len(new_a), len(new_a.columns)
            for i in range(nr):
                rs = row_status.get(i, "same")
                for j in range(nc):
                    cs = cell_status.get((i, j), "same")

                    if rs == "deleted":
                        raw = old_a.iat[i, j]
                        fill = _FILL_DELETED
                    else:
                        raw = new_a.iat[i, j]
                        fill = (
                            _FILL_ADDED   if rs == "added"   else
                            _FILL_CHANGED if cs == "changed" else
                            None
                        )

                    cell = ws.cell(i + 1, j + 1, None if raw == "" else raw)
                    cell.font = _FONT_NORMAL
                    if fill:
                        cell.fill = fill

            if nr and nc:
                _apply_border(ws, nr, nc)
                _auto_width(ws)

    # ── Summary sheet ────────────────────────────────────────────────────────
    build_summary_sheet(
        wb, ordered, new_only, deleted_only, sheet_stats,
        old_filename, new_filename
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Side-by-side Excel export
# ─────────────────────────────────────────────────────────────────────────────

def _load_source_wb(raw: bytes, filename: str):
    """Load an openpyxl Workbook from raw bytes (xlsx only, full load so all
    worksheet properties including headers/footers are available)."""
    try:
        if filename.lower().endswith(".xls"):
            return None
        return openpyxl.load_workbook(io.BytesIO(raw))
    except Exception:
        return None


def _copy_print_settings(target_ws, src_wb, sheet_name: str):
    """
    Copy page setup, margins, and header/footer from source workbook.

    page_setup and page_margins are copied attribute-by-attribute to avoid
    carrying the source sheet's _parent reference into the target, which would
    cause openpyxl to silently write headers/footers against the wrong sheet.
    HeaderFooter has no _parent so a deep copy is safe there.
    """
    if src_wb is None or sheet_name not in src_wb.sheetnames:
        return
    try:
        src_ws = src_wb[sheet_name]

        # Page setup — individual attributes only
        src_ps = src_ws.page_setup
        tgt_ps = target_ws.page_setup
        for attr in (
            'orientation', 'paperSize', 'scale', 'firstPageNumber',
            'pageOrder', 'usePrinterDefaults', 'blackAndWhite', 'draft',
            'cellComments', 'useFirstPageNumber', 'horizontalDpi',
            'verticalDpi', 'copies', 'errors',
        ):
            try:
                val = getattr(src_ps, attr, None)
                if val is not None:
                    setattr(tgt_ps, attr, val)
            except Exception:
                pass

        # Page margins — individual attributes only
        src_pm = src_ws.page_margins
        tgt_pm = target_ws.page_margins
        for attr in ('left', 'right', 'top', 'bottom', 'header', 'footer'):
            try:
                val = getattr(src_pm, attr, None)
                if val is not None:
                    setattr(tgt_pm, attr, val)
            except Exception:
                pass

        # Header / footer — safe to deep copy (no _parent reference).
        # NOTE: the attribute is "HeaderFooter" (capital H and F) in openpyxl.
        target_ws.HeaderFooter = copy.deepcopy(src_ws.HeaderFooter)

    except Exception:
        pass


def _apply_border_region(ws, r1: int, r2: int, c1: int, c2: int):
    """Apply thin borders to a rectangular region [r1..r2, c1..c2] (1-indexed, inclusive)."""
    if r2 < r1 or c2 < c1:
        return
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).border = _THIN_BORDER


def _write_sbs_header(
    ws, nc_old: int, sep_col: int, nc_new: int, old_label: str, new_label: str
):
    """Write the OLD | ◄► | NEW banner (row 1) for the side-by-side sheet."""
    # OLD header block
    if nc_old > 0:
        if nc_old > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nc_old)
        c = ws.cell(1, 1)
        c.value     = old_label
        c.fill      = _FILL_SBS_HDR_OLD
        c.font      = _FONT_SBS_HDR
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Separator cell
    c = ws.cell(1, sep_col)
    c.value     = "◄  ►"
    c.fill      = _FILL_SBS_HDR_SEP
    c.font      = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = _BORDER_SEP_DATA

    # NEW header block
    new_start = sep_col + 1
    if nc_new > 0:
        end_col = new_start + nc_new - 1
        if nc_new > 1:
            ws.merge_cells(start_row=1, start_column=new_start, end_row=1, end_column=end_col)
        c = ws.cell(1, new_start)
        c.value     = new_label
        c.fill      = _FILL_SBS_HDR_NEW
        c.font      = _FONT_SBS_HDR
        c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 26


def build_sidebyside_excel(
    old_sheets:   Dict[str, pd.DataFrame],
    new_sheets:   Dict[str, pd.DataFrame],
    new_only:     Set[str],
    deleted_only: Set[str],
    sheet_stats:  Dict[str, dict],
    sheet_data:   Dict[str, tuple],
    old_filename: str,
    new_filename: str,
    old_raw:      bytes,
    new_raw:      bytes,
) -> bytes:
    """
    Build a professional side-by-side OLD vs NEW comparison Excel workbook.

    Layout per sheet
    ────────────────
    • Data starts at row 1 — no header banner row
    • OLD data on left | narrow grey separator column | NEW data on right

    Colour coding
    ─────────────
    • Changed cell  OLD side : strikethrough red font
    • Changed cell  NEW side : yellow background  (#FFFF00)
    • Deleted row   OLD side : light-red fill
    • Added row     NEW side : light-green fill
    • Deleted sheet OLD side : pink-red fill (#FFCCCC), right side blank
    • New sheet     NEW side : light-green fill (#CCFFCC), left side blank
    • Separator column       : mid-grey fill, no borders

    Print behaviour
    ───────────────
    • No cell gridlines shown (matches source file appearance)
    • Headers / footers copied from source workbook
    • Page margins / orientation / paper size copied from source
    • fitToWidth = 1 so both OLD and NEW columns print on the same page
    """

    old_src_wb = _load_source_wb(old_raw, old_filename)
    new_src_wb = _load_source_wb(new_raw, new_filename)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheets in ascending alphabetical order (all unique names)
    all_names = sorted(set(list(old_sheets.keys()) + list(new_sheets.keys())))

    def _finalise(ws, sep_col, src_wb, src_name):
        """Apply common post-write settings: widths, no gridlines, print setup."""
        _auto_width(ws)
        ws.column_dimensions[get_column_letter(sep_col)].width = 3
        # Hide cell grid lines to match source file appearance
        ws.sheet_view.showGridLines = False
        # Copy headers / footers / margins from source
        _copy_print_settings(ws, src_wb, src_name)
        # Force landscape + both OLD and NEW columns onto one page width
        ws.page_setup.orientation = 'landscape'
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_setup.fitToWidth  = 1
        ws.page_setup.fitToHeight = 0   # unlimited rows — let it flow vertically

    for name in all_names:
        safe_name = name[:31]
        ws = wb.create_sheet(title=safe_name)

        # ── New sheet (only in revised file) ─────────────────────────────
        if name in new_only:
            ws.sheet_properties.tabColor = _TAB_COLOR["new"]
            df = new_sheets.get(name, pd.DataFrame())
            if df.empty:
                continue
            nr        = len(df)
            nc_new    = len(df.columns)
            nc_old    = nc_new          # blank mirror on the left
            sep_col   = nc_old + 1
            new_start = sep_col + 1

            for i, row_vals in enumerate(
                dataframe_to_rows(df, index=False, header=False), start=1
            ):
                for j in range(1, nc_old + 1):          # OLD side — empty
                    ws.cell(i, j)
                ws.cell(i, sep_col).fill = _FILL_SBS_SEP_DATA   # Separator
                for jj, val in enumerate(row_vals, start=new_start):  # NEW — green
                    c = ws.cell(i, jj, None if val == "" else val)
                    c.fill = _FILL_SBS_NEW_SHEET
                    c.font = _FONT_NORMAL

            _finalise(ws, sep_col, new_src_wb, name)

        # ── Deleted sheet (only in original file) ────────────────────────
        elif name in deleted_only:
            ws.sheet_properties.tabColor = _TAB_COLOR["deleted"]
            df = old_sheets.get(name, pd.DataFrame())
            if df.empty:
                continue
            nr        = len(df)
            nc_old    = len(df.columns)
            nc_new    = nc_old          # blank mirror on the right
            sep_col   = nc_old + 1
            new_start = sep_col + 1

            for i, row_vals in enumerate(
                dataframe_to_rows(df, index=False, header=False), start=1
            ):
                for j, val in enumerate(row_vals, start=1):     # OLD side — red
                    c = ws.cell(i, j, None if val == "" else val)
                    c.fill = _FILL_SBS_DEL_SHEET
                    c.font = _FONT_NORMAL
                ws.cell(i, sep_col).fill = _FILL_SBS_SEP_DATA   # Separator
                for jj in range(new_start, new_start + nc_new): # NEW side — empty
                    ws.cell(i, jj)

            _finalise(ws, sep_col, old_src_wb, name)

        # ── Common sheet (present in both files) ─────────────────────────
        else:
            old_df = old_sheets.get(name, pd.DataFrame())
            new_df = new_sheets.get(name, pd.DataFrame())

            if name in sheet_data:
                old_a, new_a, cell_status, row_status = sheet_data[name]
            else:
                old_a, new_a, cell_status, row_status, _ = compare_dataframes(old_df, new_df)

            if old_a.empty and new_a.empty:
                continue

            nr     = max(len(old_a) if not old_a.empty else 0,
                         len(new_a) if not new_a.empty else 0)
            nc_old = len(old_a.columns) if not old_a.empty else 0
            nc_new = len(new_a.columns) if not new_a.empty else 0

            sv      = sheet_stats.get(name, {})
            has_chg = (
                sv.get("changed_cells", 0)
                + sv.get("added_rows",   0)
                + sv.get("deleted_rows", 0)
            ) > 0
            ws.sheet_properties.tabColor = (
                _TAB_COLOR["modified"] if has_chg else _TAB_COLOR["unchanged"]
            )

            sep_col   = nc_old + 1
            new_start = sep_col + 1

            for i in range(nr):
                excel_row = i + 1      # data begins at row 1, no header banner
                rs = row_status.get(i, "same")

                # OLD side ──────────────────────────────────────────────
                for j in range(nc_old):
                    val = cell_str(old_a.iat[i, j]) if i < len(old_a) else ""
                    cs  = cell_status.get((i, j), "same")
                    c   = ws.cell(excel_row, j + 1, None if val == "" else val)
                    if rs == "deleted":
                        c.fill = _FILL_DELETED
                        c.font = _FONT_NORMAL
                    elif rs == "added":
                        c.value = None
                        c.font  = _FONT_NORMAL
                    elif cs == "changed":
                        c.font = _FONT_STRIKE   # strikethrough red
                    else:
                        c.font = _FONT_NORMAL

                # Separator ─────────────────────────────────────────────
                ws.cell(excel_row, sep_col).fill = _FILL_SBS_SEP_DATA

                # NEW side ──────────────────────────────────────────────
                for j in range(nc_new):
                    val = cell_str(new_a.iat[i, j]) if i < len(new_a) else ""
                    cs  = cell_status.get((i, j), "same")
                    c   = ws.cell(excel_row, new_start + j, None if val == "" else val)
                    if rs == "added":
                        c.fill = _FILL_ADDED
                        c.font = _FONT_NORMAL
                    elif rs == "deleted":
                        c.value = None
                        c.font  = _FONT_NORMAL
                    elif cs == "changed":
                        c.fill = _FILL_SBS_CHG_NEW  # yellow
                        c.font = _FONT_NORMAL
                    else:
                        c.font = _FONT_NORMAL

            _finalise(ws, sep_col, new_src_wb, name)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Inline-diff Excel export
# ─────────────────────────────────────────────────────────────────────────────

# InlineFont styles reused across all inline-diff cells
_IF_STRIKE = InlineFont(strike=True, color="C00000")   # old value — red strikethrough
_IF_NORMAL = InlineFont()                               # new / unchanged value


def _rich_changed(old_val: str, new_val: str):
    """
    Return a CellRichText showing:  ~~old~~  new
    Falls back to plain strings when one side is empty.
    """
    if old_val and new_val:
        return CellRichText([
            TextBlock(_IF_STRIKE, old_val),
            "  ",
            TextBlock(_IF_NORMAL, new_val),
        ])
    if old_val:
        return CellRichText([TextBlock(_IF_STRIKE, old_val)])
    return new_val or None


def build_inline_excel(
    old_sheets:   Dict[str, pd.DataFrame],
    new_sheets:   Dict[str, pd.DataFrame],
    new_only:     Set[str],
    deleted_only: Set[str],
    sheet_stats:  Dict[str, dict],
    sheet_data:   Dict[str, tuple],
    old_filename: str,
    new_filename: str,
    old_raw:      bytes,
    new_raw:      bytes,
) -> bytes:
    """
    Build an inline-diff Excel workbook.

    Each sheet is a single table — no side-by-side doubling.

    Cell rendering
    ──────────────
    • Changed cell     : ~~old_value~~  new_value  (rich text in one cell)
    • Unchanged cell   : value (no decoration)
    • Deleted row      : every cell strikethrough
    • Added row        : every cell, light-green fill
    • Deleted sheet    : all values strikethrough, red tab
    • New sheet        : all values, green fill, green tab

    Print behaviour
    ───────────────
    • Landscape orientation, fitToWidth = 1
    • Headers / footers / margins copied from source workbook
    • No gridlines shown
    """
    old_src_wb = _load_source_wb(old_raw, old_filename)
    new_src_wb = _load_source_wb(new_raw, new_filename)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    all_names = sorted(set(list(old_sheets.keys()) + list(new_sheets.keys())))

    def _finalise_inline(ws, src_wb, src_name):
        _auto_width(ws)
        ws.sheet_view.showGridLines = False
        _copy_print_settings(ws, src_wb, src_name)
        ws.page_setup.orientation = "landscape"
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_setup.fitToWidth  = 1
        ws.page_setup.fitToHeight = 0

    for name in all_names:
        ws = wb.create_sheet(title=name[:31])

        # ── New sheet ─────────────────────────────────────────────────────
        if name in new_only:
            ws.sheet_properties.tabColor = _TAB_COLOR["new"]
            df = new_sheets.get(name, pd.DataFrame())
            if df.empty:
                continue
            for i, row_vals in enumerate(
                dataframe_to_rows(df, index=False, header=False), start=1
            ):
                for j, val in enumerate(row_vals, start=1):
                    c = ws.cell(i, j, None if val == "" else val)
                    c.fill = _FILL_ADDED
                    c.font = _FONT_NORMAL
            _finalise_inline(ws, new_src_wb, name)

        # ── Deleted sheet ─────────────────────────────────────────────────
        elif name in deleted_only:
            ws.sheet_properties.tabColor = _TAB_COLOR["deleted"]
            df = old_sheets.get(name, pd.DataFrame())
            if df.empty:
                continue
            for i, row_vals in enumerate(
                dataframe_to_rows(df, index=False, header=False), start=1
            ):
                for j, val in enumerate(row_vals, start=1):
                    c = ws.cell(i, j, None if val == "" else val)
                    c.font = _FONT_STRIKE
            _finalise_inline(ws, old_src_wb, name)

        # ── Common sheet ──────────────────────────────────────────────────
        else:
            old_df = old_sheets.get(name, pd.DataFrame())
            new_df = new_sheets.get(name, pd.DataFrame())

            if name in sheet_data:
                old_a, new_a, cell_status, row_status = sheet_data[name]
            else:
                old_a, new_a, cell_status, row_status, _ = compare_dataframes(old_df, new_df)

            if old_a.empty and new_a.empty:
                continue

            nr = max(len(old_a) if not old_a.empty else 0,
                     len(new_a) if not new_a.empty else 0)
            nc = max(len(old_a.columns) if not old_a.empty else 0,
                     len(new_a.columns) if not new_a.empty else 0)

            sv      = sheet_stats.get(name, {})
            has_chg = (
                sv.get("changed_cells", 0)
                + sv.get("added_rows",   0)
                + sv.get("deleted_rows", 0)
            ) > 0
            ws.sheet_properties.tabColor = (
                _TAB_COLOR["modified"] if has_chg else _TAB_COLOR["unchanged"]
            )

            for i in range(nr):
                rs = row_status.get(i, "same")
                for j in range(nc):
                    cs      = cell_status.get((i, j), "same")
                    old_val = cell_str(old_a.iat[i, j]) if i < len(old_a) else ""
                    new_val = cell_str(new_a.iat[i, j]) if i < len(new_a) else ""
                    c       = ws.cell(i + 1, j + 1)

                    if rs == "deleted":
                        # Entire row deleted — strikethrough every cell
                        c.value = None if old_val == "" else old_val
                        c.font  = _FONT_STRIKE

                    elif rs == "added":
                        # Entire row added — green fill
                        c.value = None if new_val == "" else new_val
                        c.fill  = _FILL_ADDED
                        c.font  = _FONT_NORMAL

                    elif cs == "changed":
                        # Individual cell changed — rich text: ~~old~~  new
                        c.value = _rich_changed(old_val, new_val)
                        c.font  = _FONT_NORMAL

                    else:
                        # Unchanged
                        c.value = None if new_val == "" else new_val
                        c.font  = _FONT_NORMAL

            _finalise_inline(ws, new_src_wb, name)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Session-state initialisation
# ─────────────────────────────────────────────────────────────────────────────

if "report_bytes" not in st.session_state:
    st.session_state.report_bytes = None
if "report_filename" not in st.session_state:
    st.session_state.report_filename = None
if "sbs_report_bytes" not in st.session_state:
    st.session_state.sbs_report_bytes = None
if "sbs_report_filename" not in st.session_state:
    st.session_state.sbs_report_filename = None
if "inline_report_bytes" not in st.session_state:
    st.session_state.inline_report_bytes = None
if "inline_report_filename" not in st.session_state:
    st.session_state.inline_report_filename = None
if "last_file_ids" not in st.session_state:
    st.session_state.last_file_ids = (None, None)


# ─────────────────────────────────────────────────────────────────────────────
# UI — Header
# ─────────────────────────────────────────────────────────────────────────────

st.markdown(
    """
    <div class="hero">
      <h1>📊 Excel Comparator Pro</h1>
      <p>Upload two Excel workbooks — get instant, cell-level diff across every sheet</p>
    </div>
    """,
    unsafe_allow_html=True,
)

# ── Upload row ────────────────────────────────────────────────────────────────
up_col1, up_col2 = st.columns(2)

with up_col1:
    st.markdown(
        '<span class="upload-label">📁 Original File &nbsp;(Old / Baseline)</span>',
        unsafe_allow_html=True,
    )
    old_file = st.file_uploader(
        "old_upload",
        type=["xlsx", "xls"],
        key="old_file",
        label_visibility="collapsed",
        help="The file you are comparing FROM",
    )

with up_col2:
    st.markdown(
        '<span class="upload-label">📁 Revised File &nbsp;(New / Updated)</span>',
        unsafe_allow_html=True,
    )
    new_file = st.file_uploader(
        "new_upload",
        type=["xlsx", "xls"],
        key="new_file",
        label_visibility="collapsed",
        help="The file you are comparing TO",
    )

# ── Instructions (shown only while files are missing) ────────────────────────
if not old_file or not new_file:
    st.markdown(
        """
        <div class="info-box">
          <strong>How to use Excel Comparator Pro</strong>
          <ol>
            <li>Upload your <strong>Original (Old / Baseline)</strong> Excel file on the left.</li>
            <li>Upload your <strong>Revised (New / Updated)</strong> Excel file on the right.</li>
            <li>Comparison runs automatically — no extra button required.</li>
          </ol>
          <strong>What is detected:</strong>
          <ul>
            <li>🟢 <strong>New sheets</strong> — sheet added in the revised file</li>
            <li>🔴 <strong>Deleted sheets</strong> — sheet removed in the revised file</li>
            <li>🟡 <strong>Changed cells</strong> — yellow highlight with old ↦ new value</li>
            <li>🟢 <strong>Added rows</strong> — full row highlighted green</li>
            <li>🔴 <strong>Deleted rows</strong> — full row highlighted red</li>
          </ul>
          Download a <strong>highlighted Excel report</strong> with colour-coded tabs and cells.
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
# Read both files
# ─────────────────────────────────────────────────────────────────────────────

with st.spinner("Reading workbooks…"):
    old_raw = _file_bytes(old_file)
    new_raw = _file_bytes(new_file)
    old_sheets = read_excel_sheets(old_raw, old_file.name)
    new_sheets = read_excel_sheets(new_raw, new_file.name)

if not old_sheets or not new_sheets:
    st.stop()

old_names: Set[str] = set(old_sheets)
new_names: Set[str] = set(new_sheets)

new_only     = new_names - old_names
deleted_only = old_names - new_names
common       = old_names & new_names

# Preserve natural order: old sheets first, then truly-new sheets appended
ordered: List[str] = list(old_sheets.keys()) + [
    s for s in new_sheets.keys() if s not in old_sheets
]

# ─────────────────────────────────────────────────────────────────────────────
# Pre-compute diffs for all common sheets
# ─────────────────────────────────────────────────────────────────────────────

sheet_stats: Dict[str, dict] = {}
sheet_data:  Dict[str, tuple] = {}

for sname in common:
    old_a, new_a, cs, rs, stats = compare_dataframes(
        old_sheets[sname], new_sheets[sname]
    )
    sheet_stats[sname] = stats
    sheet_data[sname]  = (old_a, new_a, cs, rs)

total_changes = sum(
    v["changed_cells"] + v["added_rows"] + v["deleted_rows"]
    for v in sheet_stats.values()
)
modified_count = sum(
    1 for v in sheet_stats.values()
    if v["changed_cells"] + v["added_rows"] + v["deleted_rows"] > 0
)

# ─────────────────────────────────────────────────────────────────────────────
# Auto-generate the highlighted Excel report (once per file pair)
# ─────────────────────────────────────────────────────────────────────────────

current_ids = (id(old_raw), id(new_raw))
if st.session_state.last_file_ids != current_ids:
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    with st.spinner("Building highlighted Excel report…"):
        st.session_state.report_bytes = build_highlighted_excel(
            old_sheets, new_sheets, ordered, new_only, deleted_only,
            sheet_stats, old_file.name, new_file.name,
        )
        st.session_state.report_filename = f"excel_diff_{ts}.xlsx"
    with st.spinner("Building side-by-side comparison Excel…"):
        st.session_state.sbs_report_bytes = build_sidebyside_excel(
            old_sheets, new_sheets, new_only, deleted_only,
            sheet_stats, sheet_data, old_file.name, new_file.name,
            old_raw, new_raw,
        )
        st.session_state.sbs_report_filename = f"excel_diff_sidebyside_{ts}.xlsx"
    with st.spinner("Building inline diff Excel…"):
        st.session_state.inline_report_bytes = build_inline_excel(
            old_sheets, new_sheets, new_only, deleted_only,
            sheet_stats, sheet_data, old_file.name, new_file.name,
            old_raw, new_raw,
        )
        st.session_state.inline_report_filename = f"excel_diff_inline_{ts}.xlsx"
    st.session_state.last_file_ids = current_ids

# ─────────────────────────────────────────────────────────────────────────────
# Summary metrics
# ─────────────────────────────────────────────────────────────────────────────

st.markdown('<hr class="divider">', unsafe_allow_html=True)
st.markdown('<p class="section-title">📈 Comparison Summary</p>', unsafe_allow_html=True)

total_sheets_seen = len(old_names | new_names)

st.markdown(
    f"""
    <div class="metric-grid">
      <div class="metric-card">
        <div class="metric-value" style="color:#0f2942">{total_sheets_seen}</div>
        <div class="metric-label">Total Sheets</div>
      </div>
      <div class="metric-card">
        <div class="metric-value" style="color:#10b981">{len(new_only)}</div>
        <div class="metric-label">Added Sheets</div>
      </div>
      <div class="metric-card">
        <div class="metric-value" style="color:#ef4444">{len(deleted_only)}</div>
        <div class="metric-label">Deleted Sheets</div>
      </div>
      <div class="metric-card">
        <div class="metric-value" style="color:#f59e0b">{modified_count}</div>
        <div class="metric-label">Modified Sheets</div>
      </div>
      <div class="metric-card">
        <div class="metric-value" style="color:#6366f1">{len(common) - modified_count}</div>
        <div class="metric-label">Unchanged Sheets</div>
      </div>
      <div class="metric-card">
        <div class="metric-value" style="color:#dc2626">{total_changes:,}</div>
        <div class="metric-label">Total Changes</div>
      </div>
    </div>
    """,
    unsafe_allow_html=True,
)

# ─────────────────────────────────────────────────────────────────────────────
# Sheet overview pills
# ─────────────────────────────────────────────────────────────────────────────

st.markdown('<hr class="divider">', unsafe_allow_html=True)
st.markdown('<p class="section-title">📋 Sheet Overview</p>', unsafe_allow_html=True)

st.markdown(
    """
    <div class="legend">
      <div class="legend-item">
        <div class="legend-dot" style="background:#10b981"></div>New sheet
      </div>
      <div class="legend-item">
        <div class="legend-dot" style="background:#ef4444"></div>Deleted sheet
      </div>
      <div class="legend-item">
        <div class="legend-dot" style="background:#f59e0b"></div>Modified sheet
      </div>
      <div class="legend-item">
        <div class="legend-dot" style="background:#c5cdd8"></div>Unchanged sheet
      </div>
    </div>
    """,
    unsafe_allow_html=True,
)

pills_html = ['<div class="sheet-pills">']
for sname in ordered:
    if sname in new_only:
        cls, icon = "pill pill-new", "＋"
    elif sname in deleted_only:
        cls, icon = "pill pill-deleted", "−"
    else:
        sv = sheet_stats.get(sname, {})
        has_chg = sv.get("changed_cells", 0) + sv.get("added_rows", 0) + sv.get("deleted_rows", 0) > 0
        cls  = "pill pill-modified" if has_chg else "pill pill-unchanged"
        icon = "~" if has_chg else "✓"

    label = f"{icon} {sname}"
    pills_html.append(f'<span class="{cls}">{_esc(label)}</span>')

pills_html.append("</div>")
st.markdown("".join(pills_html), unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# Per-sheet analysis tabs
# ─────────────────────────────────────────────────────────────────────────────

st.markdown('<hr class="divider">', unsafe_allow_html=True)
st.markdown('<p class="section-title">🔍 Sheet-by-Sheet Analysis</p>', unsafe_allow_html=True)

tab_labels: List[str] = []
for sname in ordered:
    if sname in new_only:
        tab_labels.append(f"🟢 {sname}")
    elif sname in deleted_only:
        tab_labels.append(f"🔴 {sname}")
    else:
        sv = sheet_stats.get(sname, {})
        has_chg = sv.get("changed_cells", 0) + sv.get("added_rows", 0) + sv.get("deleted_rows", 0) > 0
        tab_labels.append(f"🟡 {sname}" if has_chg else f"⚪ {sname}")

tabs = st.tabs(tab_labels)

for sname, tab in zip(ordered, tabs):
    with tab:

        # ── New sheet ──────────────────────────────────────────────────────
        if sname in new_only:
            st.success(
                f"**'{sname}'** is a **new sheet** — it exists only in the revised file."
            )
            df = new_sheets[sname]
            st.caption(f"{len(df):,} rows × {len(df.columns):,} columns")
            st.dataframe(df, use_container_width=True, height=380, hide_index=True)

        # ── Deleted sheet ──────────────────────────────────────────────────
        elif sname in deleted_only:
            st.error(
                f"**'{sname}'** was **deleted** — it exists only in the original file."
            )
            df = old_sheets[sname]
            st.caption(f"{len(df):,} rows × {len(df.columns):,} columns")
            st.dataframe(df, use_container_width=True, height=380, hide_index=True)

        # ── Common sheet ───────────────────────────────────────────────────
        else:
            old_a, new_a, cell_status, row_status = sheet_data[sname]
            sv = sheet_stats[sname]
            has_chg = sv["changed_cells"] + sv["added_rows"] + sv["deleted_rows"] > 0

            if not has_chg:
                st.info(f"✅ No changes detected in **'{sname}'**.")
                st.dataframe(
                    new_sheets[sname], use_container_width=True,
                    height=320, hide_index=True,
                )
            else:
                # Mini metrics
                m1, m2, m3, m4 = st.columns(4)
                m1.metric("Rows compared",  f"{sv['total_rows']:,}")
                m2.metric(
                    "Added rows", sv["added_rows"],
                    delta=f"+{sv['added_rows']}" if sv["added_rows"] else None,
                )
                m3.metric(
                    "Deleted rows", sv["deleted_rows"],
                    delta=f"-{sv['deleted_rows']}" if sv["deleted_rows"] else None,
                    delta_color="inverse",
                )
                m4.metric("Changed cells", f"{sv['changed_cells']:,}")

                # Legend
                st.markdown(
                    """
                    <div class="legend" style="margin-top:0.8rem">
                      <div class="legend-item">
                        <div class="legend-dot" style="background:#fffbeb;border:1.5px solid #f59e0b"></div>
                        Changed cell &nbsp;(strikethrough = old value, green = new value)
                      </div>
                      <div class="legend-item">
                        <div class="legend-dot" style="background:#ecfdf5;border:1.5px solid #10b981"></div>
                        Added row
                      </div>
                      <div class="legend-item">
                        <div class="legend-dot" style="background:#fef2f2;border:1.5px solid #ef4444"></div>
                        Deleted row
                      </div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

                html = render_diff_table(old_a, new_a, cell_status, row_status)
                st.markdown(html, unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# Export section
# ─────────────────────────────────────────────────────────────────────────────

st.markdown('<hr class="divider">', unsafe_allow_html=True)
st.markdown('<p class="section-title">💾 Export</p>', unsafe_allow_html=True)

exp_col1, exp_col2 = st.columns(2)

with exp_col1:
    st.download_button(
        label="📥 Download Highlighted Excel Report",
        data=st.session_state.report_bytes,
        file_name=st.session_state.report_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary",
        help="Excel workbook with colour-coded sheet tabs and highlighted cells",
    )

with exp_col2:
    # Summary CSV
    csv_rows = []
    for sname in ordered:
        if sname in new_only:
            csv_rows.append({
                "Sheet": sname, "Status": "Added",
                "Changed Cells": 0, "Added Rows": 0, "Deleted Rows": 0, "Changed Rows": 0,
            })
        elif sname in deleted_only:
            csv_rows.append({
                "Sheet": sname, "Status": "Deleted",
                "Changed Cells": 0, "Added Rows": 0, "Deleted Rows": 0, "Changed Rows": 0,
            })
        else:
            sv = sheet_stats[sname]
            has_chg = sv["changed_cells"] + sv["added_rows"] + sv["deleted_rows"] > 0
            csv_rows.append({
                "Sheet":         sname,
                "Status":        "Modified" if has_chg else "Unchanged",
                "Changed Cells": sv["changed_cells"],
                "Added Rows":    sv["added_rows"],
                "Deleted Rows":  sv["deleted_rows"],
                "Changed Rows":  sv["changed_rows"],
            })

    csv_bytes = pd.DataFrame(csv_rows).to_csv(index=False).encode("utf-8")
    st.download_button(
        label="📊 Download Summary CSV",
        data=csv_bytes,
        file_name=f"diff_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
        mime="text/csv",
        use_container_width=True,
        help="One row per sheet — quick overview of all changes",
    )

dl_col1, dl_col2 = st.columns(2)

with dl_col1:
    st.download_button(
        label="📋 Download Side-by-Side Comparison Excel",
        data=st.session_state.sbs_report_bytes,
        file_name=st.session_state.sbs_report_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        help=(
            "Each sheet: OLD table on the left, NEW table on the right — "
            "strikethrough = changed old value, yellow = changed new value, "
            "red tab = deleted sheet, green tab = new sheet. Sheets sorted A→Z."
        ),
    )

with dl_col2:
    st.download_button(
        label="🔀 Download Inline Diff Excel",
        data=st.session_state.inline_report_bytes,
        file_name=st.session_state.inline_report_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        help=(
            "Single table per sheet — changed cells show ~~old~~  new in one cell, "
            "deleted rows are struck through, added rows are green, "
            "deleted sheets show all values struck through, new sheets have green fill."
        ),
    )
